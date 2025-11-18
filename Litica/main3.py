    # =======================================================
    # GENERAR EXCEL DE RESUMEN (CON FORMATO PROFESIONAL)
    # =======================================================
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import collections

    print("📘 Generando archivo Excel con resumen...")

    wb = Workbook()

    # ------------------------------
    # Hoja 1: Copia del reporte CSV
    # ------------------------------
    ws1 = wb.active
    ws1.title = "reporte_copiado"

    headers = ["Archivo", "Ruta Origen", "Ruta Destino", "ID Monumento", "Estado"]
    ws1.append(headers)

    for row in resultados:
        ws1.append(list(row))

    # Estilos encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")

    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Crear tabla con estilo
    tabla = Table(displayName="ReporteCopiado", ref=f"A1:E{len(resultados)+1}")
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo_tabla
    ws1.add_table(tabla)

    # Ajuste de ancho automático
    for col in ws1.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[col_letter].width = max_len + 2

    # ------------------------------
    # Hoja 2: Resumen con estilo
    # ------------------------------
    ws2 = wb.create_sheet("resumen_ids")

    regex_exc = re.compile(r"T[1-7]_\d{5}_(\d{3}_\d{7})")

    total_imagenes = len([r for r in resultados if r[4] == "COPIADO"])
    ids_monumento = [r[3] for r in resultados if r[3] != ""]
    contador_monumento = collections.Counter(ids_monumento)

    ids_excav = []
    for nombre, origen, destino, id_m, estado in resultados:
        match = regex_exc.search(nombre)
        if match:
            ids_excav.append((id_m, match.group(1)))

    contador_excav = collections.Counter(ids_excav)

    # Títulos resumen
    ws2.append(["Resumen general"])
    ws2.append(["Total de imágenes procesadas", total_imagenes])
    ws2.append(["Total de IDs de monumento únicos", len(contador_monumento)])
    ws2.append([])

    # Tabla 1: Monumentos
    ws2.append(["ID Monumento", "Cantidad de Imágenes"])
    for id_m, count in contador_monumento.items():
        ws2.append([id_m, count])

    ws2.append([])
    ws2.append(["ID Monumento", "ID Excavación", "Cantidad de Imágenes"])
    for (id_m, id_exc), count in contador_excav.items():
        ws2.append([id_m, id_exc, count])

    # Estilos en toda la hoja
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Encabezados en azul
    for row in ws2.iter_rows(min_row=5, max_row=5):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill

    # Ajuste de ancho automático para la hoja 2
    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws2.column_dimensions[col_letter].width = max_len + 2

    # Guardar archivo
    excel_path = os.path.join(REPORT_DIR, "reporte_resumen.xlsx")
    wb.save(excel_path)

    print(f"📘 Excel con formato profesional generado en: {excel_path}")
